import pandas as pd
from openpyxl import load_workbook
import re

def normalize_company_name(name):
    """Normalize company name for better matching"""
    if pd.isna(name):
        return ""
    
    name = str(name).upper().strip()

    if 'XYZ' in name:
        if 'LIMITED' in name:
            return 'XYZ LTD'

    return name.strip()

# === Update these paths ===
source_file1 = r'PnL.xlsx'
source_file2 = r'TradeHistory.xlsx'
cleartax_template = r'taxTemplate.xlsx'
output_file = r'tax_filled.xlsx'

df2 = pd.read_excel(source_file2)
df2.columns = [col.strip() for col in df2.columns]

df1 = pd.read_excel(source_file1)
df1.columns = [col.strip() for col in df1.columns]

df1['Scrip Name Normalized'] = df1['Scrip Name'].apply(normalize_company_name)
df2['SCRIPNAME Normalized'] = df2['SCRIPNAME'].apply(normalize_company_name)

print("Source file 1 columns:", df1.columns.tolist())
print("Source file 2 columns:", df2.columns.tolist())

buys = df2[df2['SELL_BUY'] == 1].copy()
sells = df2[df2['SELL_BUY'] == 2].copy()

print(f"Number of buy transactions: {len(buys)}")
print(f"Number of sell transactions: {len(sells)}")
print(f"Number of pnL transactions: {len(df1)}")

buys.sort_values(['ISIN', 'TRADE_DATE'], inplace=True)
sells.sort_values(['ISIN', 'TRADE_DATE'], inplace=True)

consolidated_records = {}

for idx, sell in sells.iterrows():
    isin = sell['ISIN']
    scrip = sell['SCRIPNAME']
    sell_qty = sell['SQTY']
    sell_date = pd.to_datetime(sell['TRADE_DATE'])

    scrip_normalized = normalize_company_name(scrip)
    pnl_match = df1[df1['Scrip Name Normalized'] == scrip_normalized]
    if len(pnl_match) > 0:
        sell_rate = pnl_match.iloc[0]['Sell Rate']
        sell_value = pnl_match.iloc[0]['Sell Value']
        profit = pnl_match.iloc[0]['Net Realized P/L']
        buy_value = pnl_match.iloc[0]['Buy Value']
    else:
        print(f"No PnL match found for scrip: {scrip}")
        continue

    buy_matches = buys[(buys['ISIN'] == isin) & (buys['BQTY'] > 0)]
    qty_needed = sell_qty
    for bidx, buy in buy_matches.iterrows():
        available_qty = buy['BQTY']  
        if available_qty == 0:
            continue
        used_qty = min(available_qty, qty_needed)
        buy_date = pd.to_datetime(buy['TRADE_DATE'])

        if isin not in consolidated_records: 
            consolidated_records[isin] = {
                'ISIN': isin,
                'Description of shares sold': scrip,
                'Number of Shares': 0,
                'Date of Purchase (DD/MM/YYYY)': buy_date,
                'Total Purchase Value': buy_value,
                'Date of Sale (DD/MM/YYYY)': sell_date,
                'Sale Price per Share': sell_rate,
            }

        consolidated_records[isin]['Number of Shares'] += used_qty
        consolidated_records[isin]['Net capital gain (auto-calculated)'] += profit

        if buy_date < consolidated_records[isin]['Date of Purchase (DD/MM/YYYY)']:
            consolidated_records[isin]['Date of Purchase (DD/MM/YYYY)'] = buy_date
        if sell_date < consolidated_records[isin]['Date of Sale (DD/MM/YYYY)']:
            consolidated_records[isin]['Date of Sale (DD/MM/YYYY)'] = sell_date
        buys.at[bidx, 'BQTY'] -= used_qty
        qty_needed -= used_qty
        if qty_needed == 0:
            break

records = list(consolidated_records.values())

for record in records:
    record['Date of Purchase (DD/MM/YYYY)'] = record['Date of Purchase (DD/MM/YYYY)'].strftime('%d/%m/%Y')
    record['Date of Sale (DD/MM/YYYY)'] = record['Date of Sale (DD/MM/YYYY)'].strftime('%d/%m/%Y')
final_df = pd.DataFrame(records)
print(f"Total records generated: {len(records)}")

wb = load_workbook(cleartax_template)
ws = wb.active

header_row = None
for row in ws.iter_rows(min_row=1, max_row=10):
    values = [cell.value for cell in row]
    if values and 'ISIN' in values:
        header_row = row[0].row
        break

if header_row is None:
    raise Exception("Header row with 'ISIN' not found in template.")

for i, record in final_df.iterrows():
    for j, col in enumerate(final_df.columns):
        ws.cell(row=header_row + 1 + i, column=j + 1, value=record[col])

wb.save(output_file)
print(f"Filled template saved to {output_file}")